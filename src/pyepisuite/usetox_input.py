"""
USEtox Input Module

This module provides the USEtoxInput class for populating USEtox Excel templates
with data from PyEPISuite API results and exporting the results.
"""

import pandas as pd
import numpy as np
import os
from typing import Dict, List, Optional, Union, Any
from pathlib import Path
import logging

column_names_dict = {
    "CAS RN": "B",
    "Name": "C",
    "PesticideTargetClass": "D",
    "PesticideChemClass": "E",
    "MW": "F",
    "Molar volume": "G",
    "pKaChemClass": "H",
    "pKa.gain": "I",
    "pKa.loss": "J",
    "KOW": "K",
    "Koc": "L", # Organic carbon-water partition coefficient
    "KH25C": "M", # Henry's Law constant at 25°C
    "Pvap25": "N", # Vapor pressure at 25°C
    "Sol25": "O", # Solubility at 25°C
    "KDOC": "P", # partitioning coefficient between dissolved organic carbon and water
    "KpSS": "Q", # partitioning coefficient between suspended solids and water
    "KpSd": "R", # partitioning coefficient between sediment and water
    "KpSl": "S", # partitioning coefficient between soil and water
    "T1/2A": "T", # Atmospheric half-life
    "T1/2W": "U", # Water half-life
    "T1/2Sd": "V", # Sediment half-life
    "T1/2Sl": "W", # Soil half-life
    "T1/2surface": "X", # half-life on indoor surface
    "T1/2P": "Y", # Plant half-life
    "T1/2wheat": "Z", # Wheat half-life
    "T1/2rice": "AA", # Rice half-life
    "T1/2tomato": "AB", # Tomato half-life
    "T1/2apple": "AC", # Apple half-life
    "T1/2lettuce": "AD", # Lettuce half-life
    "T1/2potato": "AE", # Potato half-life
    "T1/2WWTP.biodeg": "AF" # Wastewater treatment plant biodegradation half-life
}
abraham_dict = {
    "E": "AG",
    "S": "AH",
    "A": "AI",
    "B": "AJ",
    "V": "AK",
    "L": "AL",
}
BAF_dict = {
    "BAFroot": "AG", # Bioaccumulation factor for roots
    "BAFleaf": "AH", # Bioaccumulation factor for leaves
    "BTFmeat": "AI", # Biomagnification trophic factor for meat
    "BTFmilk": "AJ", # Biomagnification trophic factor for milk
    "BAFfish": "AK"  # Bioaccumulation factor for fish
}
toxicity_dict = {
    "HC20": "AR", # Hazardous concentration for 20% of species
    "HC20soil": "AS", # Hazardous concentration for 20% of species in soil
    "Data source_inh": "AT", # Data source for inhalation toxicity
    "EC10inh,noncanc": "AU", # EC10 inhalation non-cancer
    "EC10inh,noncanc GSD2": "AV", # EC10 inhalation non-cancer GSD
    "RfC": "AW", # Reference concentration
    "Data source_ing": "AX", # Data source for ingestion toxicity
    "ED10ing,noncanc": "AY", # ED10 ingestion non-cancer
    "ED10ing,noncanc GSD2": "AZ", # ED10 ingestion non-cancer GSD
    "RfDing": "BA", # Reference dose for ingestion
    "Data source_derm": "BB", # Data source for dermal toxicity
    "ED10derm,noncanc": "BC", # ED10 dermal non-cancer
    "ED10derm,noncanc GSD2": "BD", # ED10 dermal non-cancer GSD
    "RfDderm": "BE", # Reference dose for dermal
    "Data source_inh": "BF", # Data source for inhalation toxicity
    "EC10inh,noncanc": "BG", # EC10 inhalation non-cancer
    "EC10inh,noncanc GSD2": "BH", # EC10 inhalation non-cancer GSD
    "RfC": "BI", # Reference concentration
    "Data source_ing": "BJ", # Data source for ingestion toxicity
    "ED10ing,noncanc": "BK", # ED10 ingestion non-cancer
    "ED10ing,noncanc GSD2": "BL", # ED10 ingestion non-cancer GSD
    "RfDing": "BM", # Reference dose for ingestion
    "Data source_derm": "BN", # Data source for dermal toxicity
    "ED10derm,noncanc": "BO", # ED10 dermal non-cancer
    "ED10derm,noncanc GSD2": "BP", # ED10 dermal non-cancer GSD
    "RfDderm": "BQ", # Reference dose for dermal
    "ED50inh,canc": "BR", # ED50 inhalation cancer
    "ED50ing,canc": "BS", # ED50 ingestion cancer
    "ED50derm,canc": "BT" # ED50 dermal cancer
}
flagged_indicative_dict = {
    "Dissociating or Salt": "BU", # Indicates if the substance is dissociating or a salt
    "Inorganics": "BV", # Indicates if the substance is inorganic
    "Uncertainty": "BW", # Indicates uncertainty in data
    "Surfactants": "BX", # Indicates if the substance is a surfactant
    "Organometallics": "BY", # Indicates if the substance is organometallic
    "Polymer": "BZ", # Indicates if the substance is a polymer
    "Mixture": "CA", # Indicates if the substance is a mixture
    "Ecotox EF": "CB", # Ecotoxicity effect factor
    "Cancer EF inh": "CC", # Cancer effect factor for inhalation
    "Cancer EF ing": "CD", # Cancer effect factor for ingestion
    "Non-cancer EF inh": "CE", # Non-cancer effect factor for inhalation
    "Non-cancer EF ing": "CF", # Non-cancer effect factor for ingestion
    "USEtox identification number": "CG", # USEtox identification number
    "Data source": "CH", # Data source
}

logger = logging.getLogger(__name__)


class USEtoxInput:
    """
    A class for populating USEtox Excel templates with PyEPISuite data and 
    exporting results.
    
    This class handles the mapping between PyEPISuite DataFrame columns and 
    USEtox template columns, populates the template with chemical data, and 
    exports the results to Excel format.
    """
    
    # Comprehensive mapping between PyEPISuite DataFrame columns and USEtox template columns
    EPISUITE_TO_USETOX_MAPPING = {
        # Basic chemical information
        'cas': 'CAS RN',
        'name': 'Name',
        'molecular_weight': 'MW',
        'molecular_formula': None,  # Not directly mapped but could be used for validation
        
        # Physical-chemical properties
        'log_kow_estimated': 'KOW',
        'log_koc_estimated': 'Koc', 
        'henrys_law_constant_estimated': 'KH25C',
        'vapor_pressure_estimated': 'Pvap25',
        'water_solubility_logkow_estimated': 'Sol25',
        'water_solubility_waternt_estimated': 'Sol25',  # Alternative solubility source
        'melting_point_estimated': None,  # Could be used for validation
        'boiling_point_estimated': None,  # Could be used for validation
        'log_koa_estimated': None,  # Not directly in USEtox template
        
        # Environmental fate and degradation
        'atmospheric_half_life_estimated': 'T1/2A',
        'hydrocarbon_biodegradation_rate_estimated': None,  # Could be related to T1/2W
        'aerosol_adsorption_fraction_estimated': None,  # Not directly mapped
        
        # Bioaccumulation and bioconcentration
        'bioconcentration_factor': 'BAFfish',  # BCF can be proxy for BAF
        'log_bioconcentration_factor': 'BAFfish',  # Will need conversion
        'bioaccumulation_factor': 'BAFfish',
        'log_bioaccumulation_factor': 'BAFfish',  # Will need conversion
        'biotransformation_half_life': None,  # Could relate to metabolism
        
        # Environmental persistence
        'river_half_life_hours': 'T1/2W',  # Convert hours to appropriate units
        'lake_half_life_hours': 'T1/2W',   # Convert hours to appropriate units
        'fugacity_persistence': None,  # Not directly mapped
        
        # Dermal properties (not directly in USEtox but available from EPI Suite)
        'dermal_permeability_coefficient': None,
        'dermal_absorbed_dose': None,
        'lag_time_hours': None,
    }
    
    # Excel column mapping - combines all the dictionaries from the top of the file
    EXCEL_COLUMN_MAPPING = {**column_names_dict, **abraham_dict, **BAF_dict, **toxicity_dict, **flagged_indicative_dict}
    
    # Properties that have both estimated and experimental versions
    # When experimental data is available, it takes priority over estimated
    EXPERIMENTAL_PROPERTY_PRIORITY = {
        'Sol25': ['water_solubility_experimental', 'water_solubility_logkow_estimated', 'water_solubility_waternt_estimated'],
        'Pvap25': ['vapor_pressure_experimental', 'vapor_pressure_estimated'],
        'KH25C': ['henrys_law_constant_experimental', 'henrys_law_constant_estimated'],
        'KOW': ['log_kow_experimental', 'log_kow_estimated'],
        'Koc': ['log_koc_experimental', 'log_koc_estimated'],
        'T1/2A': ['atmospheric_half_life_experimental', 'atmospheric_half_life_estimated'],
        'MW': ['molecular_weight_experimental', 'molecular_weight'],
    }
    
    # USEtox columns that need unit conversions or special handling
    UNIT_CONVERSIONS = {
        'KOW': lambda x: 10**x if pd.notna(x) else np.nan,  # Convert log KOW to KOW
        'Koc': lambda x: 10**x if pd.notna(x) else np.nan,  # Convert log Koc to Koc
        'Pvap25': lambda x: x * 133.322 if pd.notna(x) else np.nan,  # mmHg to Pa
        'Sol25': lambda x: x / 1000 if pd.notna(x) else np.nan,  # mg/L to g/L
        'T1/2W': lambda x: x / 24 if pd.notna(x) else np.nan,  # Hours to days
        'BAFfish': lambda x: 10**x if pd.notna(x) and x < 10 else x if pd.notna(x) else np.nan,  # Handle both log and linear values
    }
    
    # Data source tracking for experimental vs estimated values
    DATA_SOURCE_TRACKING = {
        'Sol25': 'Solubility source',
        'Pvap25': 'Vapor pressure source', 
        'KH25C': 'Henry constant source',
        'KOW': 'Log Kow source',
        'Koc': 'Log Koc source',
        'T1/2A': 'Atmospheric half-life source',
        'MW': 'Molecular weight source',
    }
    
    def __init__(self, template_path: Optional[str] = None):
        """
        Initialize USEtoxInput with optional template path.
        
        Args:
            template_path: Path to USEtox template Excel file. If None, uses default.
        """
        self.template_path = template_path or self._get_default_template_path()
        self.template_df = None
        self.populated_df = None
        self._load_template()
        
    def _get_default_template_path(self) -> str:
        """Get the default USEtox template path."""
        # Assumes the template is in the data/usetox3 directory relative to this module
        current_dir = Path(__file__).parent
        return str(current_dir / ".." / ".." / "data" / "usetox3" / "AA_Model_substance_data_Default.xlsx")
    
    def _load_template(self):
        """Load the USEtox template Excel file."""
        try:
            self.template_df = pd.read_excel(
                self.template_path, 
                sheet_name="Substance inputs", 
                skiprows=1
            )
            logger.info(f"Loaded USEtox template with {len(self.template_df)} rows")
        except Exception as e:
            logger.error(f"Failed to load template from {self.template_path}: {e}")
            raise
    
    def get_excel_column_letter(self, usetox_column_name: str) -> Optional[str]:
        """
        Get the Excel column letter for a given USEtox column name.
        
        Args:
            usetox_column_name: Name of the USEtox column
            
        Returns:
            Excel column letter (e.g., 'A', 'B', 'AA') or None if not found
        """
        return self.EXCEL_COLUMN_MAPPING.get(usetox_column_name)
    
    def _select_best_value(self, episuite_df: pd.DataFrame, usetox_property: str, row_idx: int) -> tuple[Any, str]:
        """
        Select the best available value for a property, prioritizing experimental over estimated data.
        
        Args:
            episuite_df: DataFrame containing PyEPISuite results
            usetox_property: USEtox property name (e.g., 'Sol25', 'KOW')
            row_idx: Row index in the episuite_df
            
        Returns:
            Tuple of (value, data_source)
        """
        # Check if this property has experimental/estimated priority rules
        if usetox_property in self.EXPERIMENTAL_PROPERTY_PRIORITY:
            priority_columns = self.EXPERIMENTAL_PROPERTY_PRIORITY[usetox_property]
            
            # Try each column in priority order
            for col_name in priority_columns:
                if col_name in episuite_df.columns:
                    value = episuite_df.iloc[row_idx][col_name]
                    if pd.notna(value) and value != "" and value != 0:
                        # Determine data source
                        if 'experimental' in col_name.lower():
                            return value, 'Experimental'
                        else:
                            return value, 'PyEPISuite Estimated'
        
        # Fallback: look for any mapped column
        for episuite_col, mapped_usetox_col in self.EPISUITE_TO_USETOX_MAPPING.items():
            if mapped_usetox_col == usetox_property and episuite_col in episuite_df.columns:
                value = episuite_df.iloc[row_idx][episuite_col]
                if pd.notna(value):
                    return value, 'PyEPISuite Estimated'
        
        return np.nan, 'Not Available'
    
    def populate_from_episuite_dataframe(self, 
                                       episuite_df: pd.DataFrame,
                                       start_row: int = 0,
                                       overwrite: bool = True) -> pd.DataFrame:
        """
        Populate USEtox template with data from PyEPISuite DataFrame.
        
        Args:
            episuite_df: DataFrame containing PyEPISuite results
            start_row: Row index to start populating data (0-based)
            overwrite: Whether to overwrite existing template data
            
        Returns:
            DataFrame with populated USEtox data
        """
        if overwrite or self.populated_df is None:
            # Start with a copy of the template
            self.populated_df = self.template_df.copy()
        
        # Ensure we have enough rows in the template
        needed_rows = start_row + len(episuite_df)
        if needed_rows > len(self.populated_df):
            # Add empty rows to accommodate new data
            empty_rows = pd.DataFrame(
                index=range(len(self.populated_df), needed_rows),
                columns=self.populated_df.columns
            )
            # Fill with appropriate data types to avoid warnings
            for col in empty_rows.columns:
                if col in ['RowNr', 'MW', 'KOW', 'Koc', 'KH25C', 'Pvap25', 'Sol25']:
                    empty_rows[col] = empty_rows[col].astype('float64')
                else:
                    empty_rows[col] = empty_rows[col].astype('object')
            self.populated_df = pd.concat([self.populated_df, empty_rows], ignore_index=True)
        
        # Populate RowNr column
        for i, idx in enumerate(range(start_row, start_row + len(episuite_df))):
            self.populated_df.loc[idx, 'RowNr'] = idx + 1
        
        # Map and populate data from PyEPISuite DataFrame with experimental data priority
        data_sources = {}  # Track data sources for each chemical
        
        for row_idx in range(len(episuite_df)):
            template_row_idx = start_row + row_idx
            chemical_sources = {}
            
            # Get all possible USEtox properties that could be populated
            all_usetox_properties = set(self.EPISUITE_TO_USETOX_MAPPING.values())
            all_usetox_properties = {prop for prop in all_usetox_properties if prop is not None}
            
            for usetox_property in all_usetox_properties:
                if usetox_property in self.populated_df.columns:
                    # Get the best available value (experimental > estimated)
                    value, source = self._select_best_value(episuite_df, usetox_property, row_idx)
                    
                    if pd.notna(value):
                        # Apply unit conversions if needed
                        if usetox_property in self.UNIT_CONVERSIONS:
                            converted_value = self.UNIT_CONVERSIONS[usetox_property](value)
                        else:
                            converted_value = value
                        
                        # Ensure proper data type handling
                        if usetox_property in ['RowNr', 'MW', 'KOW', 'Koc', 'KH25C', 'Pvap25', 'Sol25', 'T1/2A', 'T1/2W', 'BAFfish']:
                            self.populated_df[usetox_property] = self.populated_df[usetox_property].astype('float64')
                        else:
                            self.populated_df[usetox_property] = self.populated_df[usetox_property].astype('object')
                        
                        # Populate the value
                        self.populated_df.loc[template_row_idx, usetox_property] = converted_value
                        chemical_sources[usetox_property] = source
                        
                        # Get Excel column letter for logging
                        excel_col = self.get_excel_column_letter(usetox_property)
                        logger.info(f"Populated {usetox_property} (Excel col {excel_col}) with {source} value: {converted_value}")
            
            # Store data sources for this chemical
            cas_number = episuite_df.iloc[row_idx].get('cas', f'Chemical_{row_idx}')
            data_sources[cas_number] = chemical_sources
        
        # Create comprehensive data source description
        source_descriptions = []
        for cas, sources in data_sources.items():
            experimental_count = sum(1 for s in sources.values() if 'Experimental' in s)
            estimated_count = sum(1 for s in sources.values() if 'Estimated' in s)
            if experimental_count > 0 and estimated_count > 0:
                source_descriptions.append(f"Mixed (Exp: {experimental_count}, Est: {estimated_count})")
            elif experimental_count > 0:
                source_descriptions.append("Experimental")
            else:
                source_descriptions.append("PyEPISuite Estimated")
        
        # Add comprehensive data source information
        self.populated_df['Data source'] = self.populated_df['Data source'].astype('object')
        for i, source_desc in enumerate(source_descriptions):
            self.populated_df.loc[start_row + i, 'Data source'] = source_desc
        
        return self.populated_df
    
    def get_excel_column_mapping_summary(self) -> Dict[str, str]:
        """
        Get a summary of all Excel column mappings for populated properties.
        
        Returns:
            Dictionary mapping USEtox property names to Excel column letters
        """
        if self.populated_df is None:
            return {}
        
        mapping_summary = {}
        for col_name in self.populated_df.columns:
            excel_col = self.get_excel_column_letter(col_name)
            if excel_col:
                mapping_summary[col_name] = excel_col
        
        return mapping_summary
    
    def print_column_mapping(self):
        """Print a formatted summary of column mappings."""
        mapping = self.get_excel_column_mapping_summary()
        if not mapping:
            print("No column mappings available. Populate the template first.")
            return
        
        print("\n📊 USEtox Column to Excel Column Mapping:")
        print("=" * 50)
        for usetox_col, excel_col in sorted(mapping.items()):
            print(f"{usetox_col:<25} → Column {excel_col}")
        print("=" * 50)
    
    def populate_from_experimental_data(self,
                                      experimental_data: Dict[str, Any],
                                      cas_column: str = 'CAS RN') -> pd.DataFrame:
        """
        Populate template with experimental data for specific chemicals.
        
        Args:
            experimental_data: Dictionary mapping CAS numbers to experimental values
            cas_column: Column name containing CAS numbers
            
        Returns:
            Updated DataFrame with experimental data
        """
        if self.populated_df is None:
            self.populated_df = self.template_df.copy()
        
        for cas, data_dict in experimental_data.items():
            # Find rows with matching CAS number
            mask = self.populated_df[cas_column] == cas
            
            if mask.any():
                # Update with experimental data
                for property_name, value in data_dict.items():
                    if property_name in self.populated_df.columns:
                        self.populated_df.loc[mask, property_name] = value
                        logger.info(f"Updated {property_name} for CAS {cas} with experimental value: {value}")
        
        return self.populated_df
    
    def add_chemical_manually(self,
                            cas: str,
                            name: str,
                            properties: Dict[str, Any],
                            row_index: Optional[int] = None) -> int:
        """
        Manually add a chemical with its properties to the template.
        
        Args:
            cas: CAS number of the chemical
            name: Name of the chemical
            properties: Dictionary of property values
            row_index: Specific row index to use (if None, appends to end)
            
        Returns:
            Row index where the chemical was added
        """
        if self.populated_df is None:
            self.populated_df = self.template_df.copy()
        
        if row_index is None:
            # Find first empty row or append
            non_empty_mask = self.populated_df['CAS RN'].notna()
            if non_empty_mask.any():
                row_index = non_empty_mask.sum()
            else:
                row_index = 0
        
        # Ensure we have enough rows
        if row_index >= len(self.populated_df):
            empty_rows = pd.DataFrame(
                index=range(len(self.populated_df), row_index + 1),
                columns=self.populated_df.columns
            )
            self.populated_df = pd.concat([self.populated_df, empty_rows], ignore_index=True)
        
        # Add basic information
        self.populated_df.loc[row_index, 'RowNr'] = row_index + 1
        # Ensure proper data types for object columns
        self.populated_df['CAS RN'] = self.populated_df['CAS RN'].astype('object')
        self.populated_df['Name'] = self.populated_df['Name'].astype('object')
        self.populated_df.loc[row_index, 'CAS RN'] = cas
        self.populated_df.loc[row_index, 'Name'] = name
        
        # Add properties
        for prop, value in properties.items():
            if prop in self.populated_df.columns:
                self.populated_df.loc[row_index, prop] = value
        
        logger.info(f"Added chemical {name} (CAS: {cas}) at row {row_index}")
        return row_index
    
    def export_to_excel(self,
                       output_path: str,
                       sheet_name: str = "Substance inputs",
                       include_headers: bool = True,
                       include_original_template: bool = False) -> None:
        """
        Export the populated USEtox data to an Excel file.
        
        Args:
            output_path: Path for the output Excel file
            sheet_name: Name of the sheet to create
            include_headers: Whether to include column headers
            include_original_template: Whether to include original template as separate sheet
        """
        if self.populated_df is None:
            raise ValueError("No data to export. Populate the template first.")
        
        # Create Excel writer
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Write main populated data
            start_row = 1 if include_headers else 0
            self.populated_df.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
                startrow=start_row,
                header=include_headers
            )
            
            # Add headers row if needed (for USEtox format compatibility)
            if include_headers:
                workbook = writer.book
                worksheet = workbook[sheet_name]
                
                # Add a title row at the top
                worksheet.insert_rows(1)
                worksheet['A1'] = 'USEtox Model Input Data - Generated by PyEPISuite'
                
                # Style the header
                from openpyxl.styles import Font, PatternFill
                header_font = Font(bold=True)
                header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
                
                for col in range(1, len(self.populated_df.columns) + 1):
                    cell = worksheet.cell(row=2, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
            
            # Include original template if requested
            if include_original_template:
                self.template_df.to_excel(
                    writer,
                    sheet_name="Original Template",
                    index=False
                )
        
        logger.info(f"Exported USEtox data to {output_path}")
    
    def get_summary_statistics(self) -> Dict[str, Any]:
        """
        Get summary statistics of the populated data.
        
        Returns:
            Dictionary containing summary statistics
        """
        if self.populated_df is None:
            return {}
        
        # Count non-empty rows
        non_empty_cas = self.populated_df['CAS RN'].notna().sum()
        
        # Get statistics for key properties
        numeric_columns = ['MW', 'KOW', 'Koc', 'KH25C', 'Pvap25', 'Sol25']
        stats = {}
        
        for col in numeric_columns:
            if col in self.populated_df.columns:
                data = pd.to_numeric(self.populated_df[col], errors='coerce')
                stats[col] = {
                    'count': data.notna().sum(),
                    'mean': data.mean(),
                    'std': data.std(),
                    'min': data.min(),
                    'max': data.max()
                }
        
        return {
            'total_chemicals': non_empty_cas,
            'property_statistics': stats,
            'data_sources': self.populated_df['Data source'].value_counts().to_dict(),
            'excel_column_mapping': self.get_excel_column_mapping_summary()
        }
    
    def get_data_source_analysis(self) -> Dict[str, Any]:
        """
        Analyze the data sources used for each property.
        
        Returns:
            Dictionary containing analysis of experimental vs estimated data usage
        """
        if self.populated_df is None:
            return {}
        
        analysis = {
            'properties_with_experimental_data': [],
            'properties_with_only_estimated_data': [],
            'mixed_source_chemicals': [],
            'experimental_data_coverage': {}
        }
        
        # Check each property that has experimental/estimated priority
        for prop in self.EXPERIMENTAL_PROPERTY_PRIORITY.keys():
            if prop in self.populated_df.columns:
                # Count non-null values
                total_values = self.populated_df[prop].notna().sum()
                if total_values > 0:
                    # Analyze data sources for this property
                    experimental_count = 0
                    estimated_count = 0
                    
                    for idx, row in self.populated_df.iterrows():
                        if pd.notna(row[prop]):
                            source = row.get('Data source', '')
                            if 'Experimental' in str(source):
                                experimental_count += 1
                            elif 'Mixed' in str(source):
                                # For mixed sources, we'd need more detailed tracking
                                experimental_count += 0.5  # Rough estimate
                                estimated_count += 0.5
                            else:
                                estimated_count += 1
                    
                    coverage = experimental_count / total_values if total_values > 0 else 0
                    analysis['experimental_data_coverage'][prop] = {
                        'total_values': total_values,
                        'experimental_count': experimental_count,
                        'estimated_count': estimated_count,
                        'experimental_percentage': coverage * 100
                    }
                    
                    if experimental_count > 0:
                        analysis['properties_with_experimental_data'].append(prop)
                    else:
                        analysis['properties_with_only_estimated_data'].append(prop)
        
        return analysis
    
    def print_data_source_summary(self):
        """Print a formatted summary of data sources used."""
        analysis = self.get_data_source_analysis()
        
        print("\n🔬 Data Source Analysis:")
        print("=" * 60)
        
        if analysis.get('experimental_data_coverage'):
            print("\n📊 Experimental Data Coverage by Property:")
            for prop, coverage in analysis['experimental_data_coverage'].items():
                excel_col = self.get_excel_column_letter(prop)
                exp_pct = coverage['experimental_percentage']
                print(f"{prop:<15} (Col {excel_col:<3}): {exp_pct:5.1f}% experimental ({coverage['experimental_count']}/{coverage['total_values']} values)")
        
        if analysis.get('properties_with_experimental_data'):
            print(f"\n✅ Properties with experimental data: {len(analysis['properties_with_experimental_data'])}")
            for prop in analysis['properties_with_experimental_data']:
                excel_col = self.get_excel_column_letter(prop)
                print(f"   • {prop} (Excel column {excel_col})")
        
        if analysis.get('properties_with_only_estimated_data'):
            print(f"\n📈 Properties with only estimated data: {len(analysis['properties_with_only_estimated_data'])}")
            for prop in analysis['properties_with_only_estimated_data']:
                excel_col = self.get_excel_column_letter(prop)
                print(f"   • {prop} (Excel column {excel_col})")
        
        print("=" * 60)
    
    def validate_data(self) -> Dict[str, List[str]]:
        """
        Validate the populated data for common issues.
        
        Returns:
            Dictionary containing validation warnings and errors
        """
        warnings = []
        errors = []
        
        if self.populated_df is None:
            errors.append("No data loaded")
            return {'warnings': warnings, 'errors': errors}
        
        # Check for missing CAS numbers
        missing_cas = self.populated_df['CAS RN'].isna().sum()
        if missing_cas > 0:
            warnings.append(f"{missing_cas} rows missing CAS numbers")
        
        # Check for duplicate CAS numbers
        duplicates = self.populated_df['CAS RN'].duplicated().sum()
        if duplicates > 0:
            warnings.append(f"{duplicates} duplicate CAS numbers found")
        
        # Check for negative molecular weights
        if 'MW' in self.populated_df.columns:
            negative_mw = (pd.to_numeric(self.populated_df['MW'], errors='coerce') < 0).sum()
            if negative_mw > 0:
                errors.append(f"{negative_mw} chemicals have negative molecular weights")
        
        # Check for unrealistic KOW values
        if 'KOW' in self.populated_df.columns:
            kow_data = pd.to_numeric(self.populated_df['KOW'], errors='coerce')
            extreme_kow = ((kow_data < 1e-10) | (kow_data > 1e10)).sum()
            if extreme_kow > 0:
                warnings.append(f"{extreme_kow} chemicals have extreme KOW values")
        
        return {'warnings': warnings, 'errors': errors}


def create_usetox_input_from_episuite(episuite_df: pd.DataFrame,
                                     output_path: str,
                                     template_path: Optional[str] = None,
                                     experimental_data: Optional[Dict] = None) -> USEtoxInput:
    """
    Convenience function to create USEtox input from PyEPISuite DataFrame.
    
    Args:
        episuite_df: DataFrame containing PyEPISuite results
        output_path: Path for output Excel file
        template_path: Optional custom template path
        experimental_data: Optional experimental data to include
        
    Returns:
        USEtoxInput instance with populated data
    """
    # Create USEtoxInput instance
    usetox_input = USEtoxInput(template_path)
    
    # Populate with PyEPISuite data
    usetox_input.populate_from_episuite_dataframe(episuite_df)
    
    # Add experimental data if provided
    if experimental_data:
        usetox_input.populate_from_experimental_data(experimental_data)
    
    # Export to Excel
    usetox_input.export_to_excel(output_path)
    
    # Print comprehensive summary
    stats = usetox_input.get_summary_statistics()
    validation = usetox_input.validate_data()
    
    print(f"✅ Successfully created USEtox input with {stats['total_chemicals']} chemicals")
    print(f"📋 Excel file exported to: {output_path}")
    
    # Show column mapping
    usetox_input.print_column_mapping()
    
    # Show data source analysis
    usetox_input.print_data_source_summary()
    
    if validation['warnings']:
        print("\n⚠️  Warnings:", validation['warnings'])
    if validation['errors']:
        print("\n❌ Errors:", validation['errors'])
    
    return usetox_input
